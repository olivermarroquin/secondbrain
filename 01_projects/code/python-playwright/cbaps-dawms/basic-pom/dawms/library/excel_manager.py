"""
ExcelManager: Utility for reading Excel test data
Supports XLSX and XLS formats for DAWMS data-driven tests
"""

import logging
from typing import List, Tuple
import openpyxl
from pathlib import Path

logger = logging.getLogger(__name__)


class ExcelManager:
    """
    ExcelManager reads Excel test data for data-driven tests.
    This is the Python equivalent of the Java ExcelManager.
    """
    
    def __init__(self, excel_file: str, sheet_name: str):
        """
        Initialize ExcelManager with file path and sheet name
        
        Args:
            excel_file: Path to Excel file
            sheet_name: Name of the sheet to read
        """
        try:
            self.file_path = Path(excel_file).resolve()
            logger.info(f"Reading Excel File ---> {self.file_path}")
            
            self.workbook = openpyxl.load_workbook(self.file_path)
            self.sheet = self.workbook[sheet_name]
            
        except Exception as e:
            logger.error(f"Error initializing ExcelManager: {e}")
            raise
    
    def read_cell(self, row_index: int, col_index: int) -> str:
        """
        Read data from a specific cell
        
        Args:
            row_index: Row index (0-based)
            col_index: Column index (0-based)
            
        Returns:
            Cell data as string
        """
        try:
            # openpyxl uses 1-based indexing
            cell_value = self.sheet.cell(row_index + 1, col_index + 1).value
            return str(cell_value) if cell_value is not None else ""
        except Exception as e:
            logger.error(f"Error reading cell [{row_index}, {col_index}]: {e}")
            return ""
    
    def get_excel_data(self) -> List[Tuple]:
        """
        Get all data from Excel sheet (excluding header row)
        
        Returns:
            List of tuples containing row data
        """
        data = []
        try:
            # Get all rows except header
            rows = list(self.sheet.iter_rows(min_row=2, values_only=True))
            
            for row_idx, row in enumerate(rows, start=2):
                # Convert row to list of strings
                row_data = tuple(str(cell) if cell is not None else "" for cell in row)
                data.append(row_data)
                
                logger.info(f"Row {row_idx}: {row_data}")
            
            return data
            
        except Exception as e:
            logger.error(f"Error reading Excel data: {e}")
            return []
        finally:
            self.close()
    
    def close(self):
        """Close workbook"""
        try:
            if self.workbook:
                self.workbook.close()
        except Exception as e:
            logger.error(f"Error closing workbook: {e}")


# Example usage:
if __name__ == "__main__":
    excel = ExcelManager(
        "src/test/resources/testdata/CalculatorTestData.xlsx",
        "MortgageData1"
    )
    data = excel.get_excel_data()
    print(f"Loaded {len(data)} rows of data")
